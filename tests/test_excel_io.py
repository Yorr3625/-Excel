from types import SimpleNamespace

import pytest
from openpyxl import Workbook

from modules import excel_io


def test_is_excel_file_accepts_all_supported_extensions_case_insensitively():
    for suffix in (".xlsx", ".xlsm", ".xlsb", ".xls", ".xltx", ".xltm"):
        assert excel_io.is_excel_file(f"заказ{suffix.upper()}")

    assert not excel_io.is_excel_file("заказ.csv")
    assert not excel_io.is_excel_file("заказ")


def test_excel_glob_pattern_contains_all_supported_extensions():
    pattern = excel_io.excel_glob_pattern()

    assert all(f"*{suffix}" in pattern for suffix in excel_io.EXCEL_SUFFIXES)


def test_load_workbook_keeps_openxml_workbook(tmp_path):
    path = tmp_path / "заказ.xlsx"
    source = Workbook()
    source.active["A1"] = "Товар"
    source.save(path)

    loaded = excel_io.load_workbook(path)

    assert loaded.active["A1"].value == "Товар"
    assert loaded.template is False


def test_load_workbook_normalises_template_formats(monkeypatch, tmp_path):
    loaded_paths = []

    def fake_load(path):
        loaded_paths.append(path)
        workbook = Workbook()
        workbook.template = True
        return workbook

    monkeypatch.setattr(excel_io, "openpyxl_load_workbook", fake_load)

    for suffix in (".xlsx", ".xlsm"):
        loaded = excel_io.load_workbook(tmp_path / f"заказ{suffix}")
        assert loaded.template is True

    for suffix in (".xltx", ".xltm"):
        loaded = excel_io.load_workbook(tmp_path / f"заказ{suffix}")
        assert loaded.template is False

    assert len(loaded_paths) == 4


def test_load_workbook_converts_binary_book_and_preserves_leading_empty_area(
    tmp_path, monkeypatch
):
    path = tmp_path / "заказ.xlsb"
    path.write_bytes(b"binary fixture")
    calls = []

    class FakeSheet:
        def to_python(self, skip_empty_area=True, nrows=None):
            calls.append((skip_empty_area, nrows))
            return [
                [None, None, None, "фм 10"],
                ["служебная строка", None, None, None],
                ["Товар A", None, None, 4],
            ]

    class FakeBook:
        sheet_names = ["Заказ", "Вторая"]

        def get_sheet_by_name(self, name):
            return FakeSheet()

        def close(self):
            calls.append("closed")

    monkeypatch.setitem(
        __import__("sys").modules,
        "python_calamine",
        SimpleNamespace(load_workbook=lambda received: FakeBook()),
    )

    loaded = excel_io.load_workbook(path)

    assert loaded.sheetnames == ["Заказ", "Вторая"]
    assert loaded.active["D1"].value == "фм 10"
    assert loaded.active["A3"].value == "Товар A"
    assert loaded.active["D3"].value == 4
    assert calls.count((False, None)) == 2
    assert calls[-1] == "closed"


def test_load_workbook_rejects_unknown_extension(tmp_path):
    path = tmp_path / "заказ.csv"
    path.write_text("Товар", encoding="utf-8")

    with pytest.raises(ValueError, match="Неподдерживаемый формат Excel"):
        excel_io.load_workbook(path)
