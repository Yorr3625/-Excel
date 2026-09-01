import pytest
from openpyxl import Workbook

from modules import paths, route_drivers
from modules.route_sheets import create_route_sheets, add_sum_column_to_all_sheets


@pytest.fixture(autouse=True)
def _isolated_route_drivers(tmp_path, monkeypatch):
    """Уводит файл водителей во временную папку, чтобы тесты не трогали config/."""

    target = tmp_path / "route_drivers.json"
    monkeypatch.setattr(paths, "ROUTE_DRIVERS_FILE", target)
    monkeypatch.setattr(route_drivers, "ROUTE_DRIVERS_FILE", target)


def _build_order_sheet():
    wb = Workbook()
    ws = wb.active
    ws.append(["Товар", "фм 4", "фм 42"])
    ws.append(["Товар A", 5, 3])
    ws.append(["Товар B", 0, 2])
    return wb, ws


def test_create_route_sheets_keeps_only_relevant_columns(groups):
    wb, ws = _build_order_sheet()

    create_route_sheets(wb, ws, groups)

    sheet1 = wb["Маршрут №1"]
    assert [c.value for c in sheet1[1]] == ["Товар", "фм 4"]
    assert [c.value for c in sheet1[2]] == ["Товар A", 5]

    sheet2 = wb["Маршрут №2"]
    assert [c.value for c in sheet2[1]] == ["Товар", "фм 42"]
    assert [c.value for c in sheet2[2]] == ["Товар A", 3]

    # исходный лист не тронут
    assert [c.value for c in ws[1]] == ["Товар", "фм 4", "фм 42"]


def test_create_route_sheets_sets_page_header_with_driver_and_area(groups):
    wb, ws = _build_order_sheet()
    route_drivers.save_route_drivers({"route_1": "Игорь"})

    create_route_sheets(wb, ws, groups)

    assert wb["Маршрут №1"].oddHeader.center.text == "Маршрут №1 (Игорь) - Текстильщик"
    assert wb["Маршрут №2"].oddHeader.center.text == "Маршрут №2 (Азер) - Центр"


def test_create_route_sheets_skips_header_for_unrecognized_group_name():
    wb, ws = _build_order_sheet()
    unnamed_group = [{"name": "Служебный лист", "names": ["фм 4"], "fill": None}]

    create_route_sheets(wb, ws, unnamed_group)

    assert wb["Служебный лист"].oddHeader.center.text is None


def test_add_sum_column_inserts_sum_formula_and_totals():
    wb = Workbook()
    ws = wb.active
    ws.title = "Маршрут №1"
    ws.append(["Товар", "фм 4"])
    ws.append(["Товар A", 5])
    ws.append(["Товар B", 3])

    route_totals = add_sum_column_to_all_sheets(wb)

    assert ws.cell(row=1, column=2).value == "Сумма"
    assert ws.cell(row=2, column=2).value == "=SUM(C2:C2)"
    assert ws.cell(row=3, column=2).value == "=SUM(C3:C3)"
    assert route_totals["Маршрут №1"] == 8


def test_add_sum_column_deletes_rows_without_orders():
    wb = Workbook()
    ws = wb.active
    ws.title = "Маршрут №1"
    ws.append(["Товар", "фм 4"])
    ws.append(["Товар A", 5])
    ws.append(["Товар B (нет заказа)", None])

    add_sum_column_to_all_sheets(wb)

    values = [row[0].value for row in ws.iter_rows(min_row=2)]
    assert "Товар B (нет заказа)" not in values


def test_add_sum_column_writes_total_row():
    wb = Workbook()
    ws = wb.active
    ws.title = "Маршрут №1"
    ws.append(["Товар", "фм 4"])
    ws.append(["Товар A", 5])
    ws.append(["Товар B", 3])

    add_sum_column_to_all_sheets(wb)

    total_row = ws.max_row
    assert ws.cell(total_row, 1).value == "ИТОГО"
    assert ws.cell(total_row, 2).value == "=SUM(B2:B3)"


def test_add_sum_column_only_reports_totals_for_route_sheets():
    wb = Workbook()
    ws = wb.active
    ws.title = "Служебный лист"
    ws.append(["Товар", "фм 4"])
    ws.append(["Товар A", 5])

    route_totals = add_sum_column_to_all_sheets(wb)

    assert route_totals == {}
