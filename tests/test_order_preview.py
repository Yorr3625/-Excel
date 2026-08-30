from openpyxl import Workbook
import pytest

from modules.order_preview import PreviewError, build_order_preview
from modules.reporter import format_preview
from modules.styles import conflict_fill


def test_build_order_preview_matches_pipeline_totals(
    tmp_path, groups, sample_order_workbook
):
    input_file = tmp_path / "заказ.xlsx"
    sample_order_workbook.save(input_file)
    source_before = input_file.read_bytes()

    preview = build_order_preview(str(input_file), groups, conflict_fill, "Область")

    assert preview["file_name"] == "заказ.xlsx"
    assert preview["file_extension"] == ".xlsx"
    assert preview["sheet_name"] == "Sheet"
    assert preview["sheet_count"] == 1
    assert preview["mode"] == "Область"
    assert preview["document_type"] == "Похоже на Excel-заказ"
    assert preview["order_rows"] == 2
    assert preview["total_found"] == 2
    assert preview["route_totals"] == {"Маршрут №1": 5, "Маршрут №2": 5}
    assert [route["stores"] for route in preview["route_rows"]] == [
        ["фм 4"],
        ["фм 42"],
    ]
    assert preview["grand_total"] == 10
    assert [route["rows"] for route in preview["route_rows"]] == [2, 2]
    assert preview["warnings"] == []
    preview_text = format_preview(preview)
    assert "Режим: Область" in preview_text
    assert "Магазины: фм 4" in preview_text
    assert input_file.read_bytes() == source_before
    assert not (tmp_path / "data").exists()


def test_build_order_preview_reports_conflicts_and_unknown_stores(
    tmp_path, groups
):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["Товар", "служебный1", "служебный2", "фм 4 фм 42", "неизвестный магазин"])
    worksheet.append(["служебная строка", "", "", "", ""])
    worksheet.append(["Товар A", "", "", 5, 2])
    input_file = tmp_path / "конфликт.xlsx"
    workbook.save(input_file)

    preview = build_order_preview(input_file, groups, conflict_fill, "Город")

    assert preview["conflict_count"] == 1
    assert preview["conflicts"] == [
        {
            "cell": "B1",
            "text": "фм 4 фм 42",
            "routes": ["Маршрут №1", "Маршрут №2"],
        }
    ]
    assert preview["unknown_stores"] == ["неизвестный магазин"]
    assert any("конфликтующих" in warning for warning in preview["warnings"])
    assert any("неизвестные" in warning for warning in preview["warnings"])


def test_preview_uses_shared_loader_for_binary_workbook(
    tmp_path, monkeypatch, groups
):
    input_file = tmp_path / "заказ.xlsb"
    input_file.write_bytes(b"binary fixture")

    converted = Workbook()
    worksheet = converted.active
    worksheet.append(["Товар", "служебный1", "служебный2", "фм 4", "фм 42"])
    worksheet.append(["служебная строка", "", "", "", ""])
    worksheet.append(["Товар A", "", "", 5, 3])
    monkeypatch.setattr("modules.pipeline.load_workbook", lambda path: converted)

    preview = build_order_preview(input_file, groups, conflict_fill, "Область")

    assert preview["file_extension"] == ".xlsb"
    assert preview["route_totals"] == {"Маршрут №1": 5, "Маршрут №2": 3}
    assert preview["grand_total"] == 8


def test_build_order_preview_wraps_corrupt_workbook(tmp_path, groups):
    input_file = tmp_path / "повреждённый.xlsx"
    input_file.write_bytes(b"not an xlsx workbook")

    with pytest.raises(PreviewError, match="предварительный просмотр"):
        build_order_preview(input_file, groups, conflict_fill)


def test_preview_does_not_create_history_or_result_files(
    tmp_path, monkeypatch, groups, sample_order_workbook
):
    input_file = tmp_path / "заказ.xlsx"
    sample_order_workbook.save(input_file)
    monkeypatch.chdir(tmp_path)

    build_order_preview(input_file, groups, conflict_fill)

    assert not list(tmp_path.glob("**/*_processed.xlsx"))
    assert not list(tmp_path.glob("**/*.log"))
    assert not list(tmp_path.glob("processed_files.json"))
    assert not list(tmp_path.glob("volume_history.json"))
