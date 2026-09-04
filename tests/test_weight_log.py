import pytest

from modules import paths, weight_log


@pytest.fixture(autouse=True)
def _isolated_data_dir(tmp_path, monkeypatch):
    """Уводит книгу веса во временную папку, чтобы тесты не трогали data/."""

    target = tmp_path / "weight_log.xlsx"
    monkeypatch.setattr(paths, "WEIGHT_LOG_FILE", target)
    monkeypatch.setattr(weight_log, "WEIGHT_LOG_FILE", target)


def test_load_returns_empty_when_file_is_missing():
    assert weight_log.load_weight_rows() == []


def test_add_without_exact_weight_computes_total_from_average():
    entry = weight_log.add_weight_row("Товар А", box_count=10, avg_weight=2.5, exact_weight=None)

    assert entry["total"] == 25.0
    assert weight_log.load_weight_rows()[0]["total"] == 25.0


def test_exact_weight_is_gross_and_total_is_net_after_subtracting_boxes():
    """Если позицию взвесили с ящиками — итог = грязный вес минус вес ящиков (кол-во × средний)."""

    entry = weight_log.add_weight_row("Товар Б", box_count=10, avg_weight=0.5, exact_weight=30.0)

    assert entry["total"] == 25.0
    assert weight_log.load_weight_rows()[0]["exact_weight"] == 30.0
    assert weight_log.load_weight_rows()[0]["total"] == 25.0


def test_rows_keep_insertion_order():
    weight_log.add_weight_row("Первый", box_count=1, avg_weight=1, exact_weight=None)
    weight_log.add_weight_row("Второй", box_count=1, avg_weight=1, exact_weight=None)

    names = [row["name"] for row in weight_log.load_weight_rows()]
    assert names == ["Первый", "Второй"]


def test_delete_weight_row_removes_only_matching_id():
    first = weight_log.add_weight_row("Первый", box_count=1, avg_weight=1, exact_weight=None)
    second = weight_log.add_weight_row("Второй", box_count=1, avg_weight=1, exact_weight=None)

    weight_log.delete_weight_row(first["id"])

    rows = weight_log.load_weight_rows()
    assert len(rows) == 1
    assert rows[0]["id"] == second["id"]


def test_delete_missing_id_does_nothing():
    weight_log.add_weight_row("Единственный", box_count=1, avg_weight=1, exact_weight=None)

    weight_log.delete_weight_row("не-существующий-id")

    assert len(weight_log.load_weight_rows()) == 1


def test_add_weight_row_creates_parent_directory(tmp_path, monkeypatch):
    target = tmp_path / "нет-такой-папки" / "weight_log.xlsx"
    monkeypatch.setattr(weight_log, "WEIGHT_LOG_FILE", target)

    weight_log.add_weight_row("Товар", box_count=1, avg_weight=1, exact_weight=None)

    assert target.exists()


def test_order_and_route_binding_are_saved():
    entry = weight_log.add_weight_row(
        "Товар",
        box_count=1,
        avg_weight=1,
        exact_weight=None,
        order_file="заказ.xlsx",
        route="Маршрут №2",
    )

    assert entry["order_file"] == "заказ.xlsx"
    assert entry["route"] == "Маршрут №2"

    row = weight_log.load_weight_rows()[0]
    assert row["order_file"] == "заказ.xlsx"
    assert row["route"] == "Маршрут №2"


def test_binding_defaults_to_empty_string():
    entry = weight_log.add_weight_row("Товар", box_count=1, avg_weight=1, exact_weight=None)

    assert entry["order_file"] == ""
    assert entry["route"] == ""
    assert weight_log.load_weight_rows()[0]["order_file"] == ""


def test_update_weight_row_recomputes_total_and_keeps_date_and_id():
    entry = weight_log.add_weight_row("Черновик", box_count=10, avg_weight=1, exact_weight=None)

    updated = weight_log.update_weight_row(
        entry["id"],
        "Исправлено",
        box_count=4,
        avg_weight=2.5,
        exact_weight=None,
        order_file="заказ.xlsx",
        route="Маршрут №1",
    )

    assert updated["id"] == entry["id"]
    assert updated["date"] == entry["date"]
    assert updated["name"] == "Исправлено"
    assert updated["total"] == 10.0
    assert updated["order_file"] == "заказ.xlsx"

    row = weight_log.load_weight_rows()[0]
    assert row["name"] == "Исправлено"
    assert row["total"] == 10.0
    assert row["route"] == "Маршрут №1"


def test_update_weight_row_missing_id_returns_none():
    weight_log.add_weight_row("Товар", box_count=1, avg_weight=1, exact_weight=None)

    assert weight_log.update_weight_row("нет-такого-id", "Х", box_count=1, avg_weight=1, exact_weight=None) is None


def test_update_weight_row_on_missing_file_returns_none():
    assert weight_log.update_weight_row("любой-id", "Х", box_count=1, avg_weight=1, exact_weight=None) is None


def test_last_avg_weight_for_returns_most_recent_match():
    weight_log.add_weight_row("Хлеб", box_count=1, avg_weight=2.0, exact_weight=None)
    weight_log.add_weight_row("Батон", box_count=1, avg_weight=1.0, exact_weight=None)
    weight_log.add_weight_row("Хлеб", box_count=1, avg_weight=2.4, exact_weight=None)

    assert weight_log.last_avg_weight_for("Хлеб") == 2.4
    assert weight_log.last_avg_weight_for("хлеб") == 2.4
    assert weight_log.last_avg_weight_for("  Хлеб  ") == 2.4


def test_last_avg_weight_for_unknown_name_returns_none():
    weight_log.add_weight_row("Хлеб", box_count=1, avg_weight=2.0, exact_weight=None)

    assert weight_log.last_avg_weight_for("Неизвестный товар") is None
    assert weight_log.last_avg_weight_for("") is None


def test_reads_legacy_rows_without_order_and_route_columns(tmp_path, monkeypatch):
    """Книга, созданная до появления привязки, не должна ломать чтение."""

    from openpyxl import Workbook

    target = tmp_path / "legacy_weight_log.xlsx"
    monkeypatch.setattr(weight_log, "WEIGHT_LOG_FILE", target)

    workbook = Workbook()
    workbook.active.append(
        ("ID", "Дата", "Наименование", "Кол-во ящиков", "Средний вес ящика, кг", "Точный вес, кг", "Итого, кг")
    )
    workbook.active.append(("legacy-id", "2026-01-01", "Старая запись", 5, 2, None, 10))
    workbook.save(target)

    row = weight_log.load_weight_rows()[0]
    assert row["order_file"] == ""
    assert row["route"] == ""
    assert row["total"] == 10


def test_update_weight_row_on_legacy_book_adds_missing_columns(tmp_path, monkeypatch):
    """Правка старой (7-колоночной) строки не должна падать на отсутствующих ячейках."""

    from openpyxl import Workbook

    target = tmp_path / "legacy_weight_log.xlsx"
    monkeypatch.setattr(weight_log, "WEIGHT_LOG_FILE", target)

    workbook = Workbook()
    workbook.active.append(
        ("ID", "Дата", "Наименование", "Кол-во ящиков", "Средний вес ящика, кг", "Точный вес, кг", "Итого, кг")
    )
    workbook.active.append(("legacy-id", "2026-01-01", "Старая запись", 5, 2, None, 10))
    workbook.save(target)

    updated = weight_log.update_weight_row(
        "legacy-id", "Обновлено", box_count=5, avg_weight=2, exact_weight=None, order_file="заказ.xlsx", route=""
    )

    assert updated["order_file"] == "заказ.xlsx"

    row = weight_log.load_weight_rows()[0]
    assert row["name"] == "Обновлено"
    assert row["order_file"] == "заказ.xlsx"


def test_stage_and_store_are_saved_and_default_to_empty():
    with_stage = weight_log.add_weight_row(
        "Огурец",
        box_count=5,
        avg_weight=2,
        exact_weight=None,
        order_file="заказ.xlsx",
        route="Маршрут №1",
        stage=weight_log.STAGE_STORE_SHIPMENT,
        store="фм 10",
    )
    without_stage = weight_log.add_weight_row("Помидор", box_count=3, avg_weight=1.5, exact_weight=None)

    assert with_stage["stage"] == weight_log.STAGE_STORE_SHIPMENT
    assert with_stage["store"] == "фм 10"
    assert without_stage["stage"] == ""
    assert without_stage["store"] == ""

    rows = weight_log.load_weight_rows()
    assert rows[0]["stage"] == weight_log.STAGE_STORE_SHIPMENT
    assert rows[0]["store"] == "фм 10"
    assert rows[1]["stage"] == ""
    assert rows[1]["store"] == ""


def test_update_weight_row_changes_stage_and_store():
    entry = weight_log.add_weight_row("Товар", box_count=1, avg_weight=1, exact_weight=None)

    updated = weight_log.update_weight_row(
        entry["id"],
        "Товар",
        box_count=1,
        avg_weight=1,
        exact_weight=None,
        order_file="",
        route="",
        stage=weight_log.STAGE_UNLOADING,
        store="",
    )

    assert updated["stage"] == weight_log.STAGE_UNLOADING
    assert weight_log.load_weight_rows()[0]["stage"] == weight_log.STAGE_UNLOADING


def test_reads_legacy_rows_without_stage_and_store_columns(tmp_path, monkeypatch):
    """Книга без колонок Этап/Магазин (до этого редизайна) не должна ломать чтение."""

    from openpyxl import Workbook

    target = tmp_path / "legacy_weight_log.xlsx"
    monkeypatch.setattr(weight_log, "WEIGHT_LOG_FILE", target)

    workbook = Workbook()
    workbook.active.append(
        (
            "ID", "Дата", "Наименование", "Кол-во ящиков", "Средний вес ящика, кг",
            "Точный вес, кг", "Итого, кг", "Заказ", "Маршрут",
        )
    )
    workbook.active.append(
        ("legacy-id", "2026-01-01", "Старая запись", 5, 2, None, 10, "заказ.xlsx", "Маршрут №1")
    )
    workbook.save(target)

    row = weight_log.load_weight_rows()[0]
    assert row["order_file"] == "заказ.xlsx"
    assert row["stage"] == ""
    assert row["store"] == ""

    updated = weight_log.update_weight_row(
        "legacy-id", "Обновлено", box_count=5, avg_weight=2, exact_weight=None,
        order_file="заказ.xlsx", route="Маршрут №1", stage=weight_log.STAGE_LOADING, store="",
    )
    assert updated["stage"] == weight_log.STAGE_LOADING


def test_known_names_for_order_returns_recent_first_without_duplicates():
    weight_log.add_weight_row("Огурец", box_count=1, avg_weight=1, exact_weight=None, order_file="заказ.xlsx")
    weight_log.add_weight_row("Помидор", box_count=1, avg_weight=1, exact_weight=None, order_file="заказ.xlsx")
    weight_log.add_weight_row("Другой заказ", box_count=1, avg_weight=1, exact_weight=None, order_file="другой.xlsx")
    weight_log.add_weight_row("огурец", box_count=1, avg_weight=1, exact_weight=None, order_file="заказ.xlsx")

    names = weight_log.known_names_for_order("заказ.xlsx")

    assert names == ["огурец", "Помидор"]


def test_known_names_for_order_empty_for_unknown_or_blank_order():
    weight_log.add_weight_row("Огурец", box_count=1, avg_weight=1, exact_weight=None, order_file="заказ.xlsx")

    assert weight_log.known_names_for_order("нет-такого-заказа.xlsx") == []
    assert weight_log.known_names_for_order("") == []
    assert weight_log.known_names_for_order("   ") == []
