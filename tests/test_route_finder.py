from openpyxl import Workbook

from modules.route_finder import find_and_mark_routes
from modules.styles import conflict_fill as CONFLICT_FILL


def test_single_route_match_fills_cell_and_counts(groups):
    wb = Workbook()
    ws = wb.active
    ws.append(["Товар", "фм 4"])

    stats = find_and_mark_routes(ws, groups, CONFLICT_FILL)

    cell = ws.cell(row=1, column=2)
    assert cell.fill.start_color.rgb == groups[0]["fill"].start_color.rgb
    assert stats["route_count"]["Маршрут №1"] == 1
    assert stats["route_count"]["Маршрут №2"] == 0
    assert stats["total_found"] == 1
    assert stats["conflict_count"] == 0


def test_word_boundary_does_not_confuse_route_1_and_route_2(groups):
    # "фм 4" не должен ложно совпасть в тексте "фм 42" (маршрут №2).
    wb = Workbook()
    ws = wb.active
    ws.append(["Товар", "фм 42"])

    stats = find_and_mark_routes(ws, groups, CONFLICT_FILL)

    assert stats["route_count"]["Маршрут №1"] == 0
    assert stats["route_count"]["Маршрут №2"] == 1


def test_conflict_when_multiple_routes_match(groups):
    wb = Workbook()
    ws = wb.active
    ws.append(["Товар", "фм 4 фм 42"])

    stats = find_and_mark_routes(ws, groups, CONFLICT_FILL)

    cell = ws.cell(row=1, column=2)
    assert cell.fill.start_color.rgb == CONFLICT_FILL.start_color.rgb
    assert stats["conflict_count"] == 1
    assert stats["conflict_list"][0]["маршруты"] == ["Маршрут №1", "Маршрут №2"]
    assert stats["total_found"] == 0


def test_unknown_store_only_detected_in_header_row(groups):
    wb = Workbook()
    ws = wb.active
    ws.append(["Товар", "неизвестный магазин"])
    ws.append(["Товар A", "неизвестный магазин"])

    stats = find_and_mark_routes(ws, groups, CONFLICT_FILL)

    # только шапка (строка 1, столбец > 1) считается неизвестным магазином
    assert stats["unknown_stores"] == ["неизвестный магазин"]


def test_case_insensitive_matching(groups):
    wb = Workbook()
    ws = wb.active
    ws.append(["Товар", "ФМ 4"])

    stats = find_and_mark_routes(ws, groups, CONFLICT_FILL)

    assert stats["route_count"]["Маршрут №1"] == 1
