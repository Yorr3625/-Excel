from openpyxl import Workbook

from modules.excel_cleaner import (
    remove_unused_rows_and_cols,
    delete_total_rows,
    delete_columns_by_text,
)


def test_remove_unused_rows_and_cols_removes_row2_and_cols_2_3():
    wb = Workbook()
    ws = wb.active

    ws.append(["A1", "B1", "C1", "D1"])
    ws.append(["A2", "B2", "C2", "D2"])
    ws.append(["A3", "B3", "C3", "D3"])

    remove_unused_rows_and_cols(ws)

    assert ws.max_row == 2
    assert [c.value for c in ws[1]] == ["A1", "D1"]
    assert [c.value for c in ws[2]] == ["A3", "D3"]


def test_delete_total_rows_removes_rows_starting_with_keywords():
    wb = Workbook()
    ws = wb.active

    ws.append(["Товар A", 1])
    ws.append(["Итого", 1])
    ws.append(["Товар B", 2])
    ws.append(["ИТОГ по маршруту", 2])
    ws.append(["Всего", 3])
    ws.append(["Товар C", 4])

    delete_total_rows(ws)

    values = [row[0].value for row in ws.iter_rows()]
    assert values == ["Товар A", "Товар B", "Товар C"]


def test_delete_total_rows_does_not_match_word_containing_keyword():
    wb = Workbook()
    ws = wb.active

    ws.append(["Итоговая позиция", 1])

    delete_total_rows(ws)

    assert ws.max_row == 1


def test_delete_columns_by_text_removes_matching_columns_case_insensitive():
    wb = Workbook()
    ws = wb.active

    ws.append(["Товар", "Ищенко БЕЗ НДС", "фм 4"])
    ws.append(["Товар A", 10, 5])

    delete_columns_by_text(ws, ["ищенко без ндс"])

    assert [c.value for c in ws[1]] == ["Товар", "фм 4"]
    assert [c.value for c in ws[2]] == ["Товар A", 5]


def test_delete_columns_by_text_no_match_leaves_sheet_untouched():
    wb = Workbook()
    ws = wb.active

    ws.append(["Товар", "фм 4"])

    delete_columns_by_text(ws, ["не найдётся"])

    assert [c.value for c in ws[1]] == ["Товар", "фм 4"]
