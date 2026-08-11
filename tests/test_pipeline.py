import os

from openpyxl import load_workbook

from modules.pipeline import process_order, detect_mode
from modules.styles import conflict_fill


def test_process_order_full_pipeline(tmp_path, monkeypatch, groups, sample_order_workbook):
    monkeypatch.chdir(tmp_path)
    input_file = tmp_path / "заказ.xlsx"
    sample_order_workbook.save(input_file)

    settings = {"open_file_after_processing": False, "open_folder_after_processing": False}

    output_file, log_file, stats = process_order(str(input_file), settings, groups, conflict_fill)

    assert os.path.isfile(output_file)
    assert os.path.isfile(log_file)
    assert stats["conflict_count"] == 0
    assert stats["route_totals"] == {"Маршрут №1": 5, "Маршрут №2": 5}

    result_wb = load_workbook(output_file)
    assert "Маршрут №1" in result_wb.sheetnames
    assert "Маршрут №2" in result_wb.sheetnames


def test_process_order_respects_open_file_setting(tmp_path, monkeypatch, groups, sample_order_workbook):
    monkeypatch.chdir(tmp_path)
    input_file = tmp_path / "заказ.xlsx"
    sample_order_workbook.save(input_file)

    calls = []
    monkeypatch.setattr(os, "startfile", lambda path: calls.append(path), raising=False)

    settings = {"open_file_after_processing": True, "open_folder_after_processing": False}
    output_file, _, _ = process_order(str(input_file), settings, groups, conflict_fill)

    assert calls == [output_file]


def test_detect_mode_picks_group_set_with_more_matches(tmp_path, groups, sample_order_workbook):
    input_file = tmp_path / "заказ.xlsx"
    sample_order_workbook.save(input_file)

    empty_groups = [{"name": "Пусто", "names": ["не найдётся нигде"], "fill": groups[0]["fill"]}]

    best_mode, scores = detect_mode(
        str(input_file),
        {"Хорошо": groups, "Плохо": empty_groups},
        conflict_fill,
    )

    assert best_mode == "Хорошо"
    assert scores["Хорошо"] > scores["Плохо"]
    assert scores["Плохо"] == 0
