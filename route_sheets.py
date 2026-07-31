from copy import copy
from openpyxl.utils import get_column_letter


def create_route_sheets(wb, ws, groups):

    for group in groups:

        sheet_name = group["name"]

        # если лист уже существует - удаляем
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

        new_ws = wb.create_sheet(sheet_name)


        # ==========================
        # КОПИРУЕМ ШАПКУ
        # ==========================

        new_ws["A1"] = "№"
        new_ws["B1"] = "Сумма"


        for cell in ws[1]:

            new_cell = new_ws.cell(
                row=1,
                column=cell.column + 2,
                value=cell.value
            )

            if cell.has_style:
                new_cell._style = copy(cell._style)



        # ==========================
        # КОПИРУЕМ СТРОКИ МАРШРУТА
        # ==========================

        new_row = 2
        number = 1


        for row in ws.iter_rows(min_row=2):

            found = False


            for cell in row:

                text = str(cell.value or "").lower()


                for name in group["names"]:

                    if name.lower() in text:

                        found = True
                        break


                if found:
                    break



            if found:


                # номер
                new_ws.cell(
                    row=new_row,
                    column=1,
                    value=number
                )


                # копируем строку
                for cell in row:

                    new_cell = new_ws.cell(
                        row=new_row,
                        column=cell.column + 2,
                        value=cell.value
                    )


                    if cell.has_style:
                        new_cell._style = copy(cell._style)



                # сумма
                last_column = new_ws.max_column


                new_ws.cell(
                    row=new_row,
                    column=2,
                    value=(
                        f"=SUM(C{new_row}:"
                        f"{get_column_letter(last_column)}{new_row})"
                    )
                )


                new_row += 1
                number += 1



        # ==========================
        # ШИРИНА КОЛОНОК
        # ==========================

        for column in ws.column_dimensions:

            if column in new_ws.column_dimensions:

                new_ws.column_dimensions[column].width = (
                    ws.column_dimensions[column].width
                )


        # закрепляем шапку
        new_ws.freeze_panes = "A2"