"""Common Excel format detection and workbook loading.

Open XML workbooks are kept in openpyxl so their styles and workbook structure
survive processing. Legacy and binary workbooks are read with python-calamine
and converted to an openpyxl workbook, which is the representation expected by
this project's processing pipeline.
"""

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl import load_workbook as openpyxl_load_workbook


OPENXML_SUFFIXES = (".xlsx", ".xlsm", ".xltx", ".xltm")
BINARY_SUFFIXES = (".xls", ".xlsb")
EXCEL_SUFFIXES = OPENXML_SUFFIXES + BINARY_SUFFIXES

# Kept as data rather than being duplicated in each UI. Browsers use the MIME
# type when available, while the extension list is still needed for filtering.
EXCEL_MIME_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": [".xlsx"],
    "application/vnd.ms-excel.sheet.macroEnabled.12": [".xlsm"],
    "application/vnd.ms-excel.sheet.binary.macroEnabled.12": [".xlsb"],
    "application/vnd.ms-excel": [".xls"],
    "application/vnd.openxmlformats-officedocument.spreadsheetml.template": [".xltx"],
    "application/vnd.ms-excel.template.macroEnabled.12": [".xltm"],
}


def is_excel_file(path: str | Path) -> bool:
    """Return whether *path* has one of the supported Excel extensions."""

    return Path(path).suffix.casefold() in EXCEL_SUFFIXES


def excel_glob_pattern() -> str:
    """Return a file-dialog pattern containing all supported extensions."""

    return " ".join(f"*{suffix}" for suffix in EXCEL_SUFFIXES)


def load_workbook(path: str | Path):
    """Load any supported Excel book as an openpyxl workbook.

    ``openpyxl`` handles the XML formats directly. ``python-calamine`` is
    imported lazily because it is only needed for the legacy/binary formats;
    this also keeps XML-only use and existing installations usable while the
    optional dependency is being installed.

    Template files are intentionally normalised to regular workbooks because
    the output writer always saves a processed ``.xlsx`` file. VBA projects
    are not copied into the result.
    """

    path = Path(path)
    suffix = path.suffix.casefold()

    if suffix in OPENXML_SUFFIXES:
        workbook = openpyxl_load_workbook(path)

        if suffix in (".xltx", ".xltm"):
            workbook.template = False

        return workbook

    if suffix in BINARY_SUFFIXES:
        return _load_binary_workbook(path)

    raise ValueError(
        f"Неподдерживаемый формат Excel: {path.suffix or 'без расширения'}"
    )


def _load_binary_workbook(path: Path):
    try:
        from python_calamine import load_workbook as calamine_load_workbook
    except ImportError as error:
        raise RuntimeError(
            "Для чтения файлов .xls и .xlsb установите зависимость "
            "python-calamine"
        ) from error

    source = calamine_load_workbook(path)

    try:
        workbook = Workbook()
        # Workbook() creates a default sheet. It is replaced by the source
        # sheets so the first sheet remains the active sheet as in the input.
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        for sheet_name in source.sheet_names:
            source_sheet = source.get_sheet_by_name(sheet_name)
            target_sheet = workbook.create_sheet(title=sheet_name)
            rows = source_sheet.to_python(skip_empty_area=False)

            for row_number, values in enumerate(rows, start=1):
                for column_number, value in enumerate(values, start=1):
                    target_sheet.cell(row=row_number, column=column_number).value = _plain_value(value)

        if not workbook.worksheets:
            workbook.create_sheet()

        workbook.active = 0
        return workbook
    finally:
        close = getattr(source, "close", None)
        if callable(close):
            close()


def _plain_value(value: Any) -> Any:
    """Convert wrapper values, if a Calamine version returns any, to Python."""

    # Current python-calamine's to_python() already returns regular Python
    # values. Keep this deliberately conservative: dates, decimals and numbers
    # should reach openpyxl unchanged.
    return value
