"""Учёт веса позиций: наименование, ящики, средний и точный вес.

Хранится отдельной книгой data/weight_log.xlsx, а не в файлах заказов —
записи не связаны с конкретной обработкой и их удобно открыть и поправить
вручную в Excel. Колонка ID скрыта от пользователя и нужна только для того,
чтобы удалить конкретную строку, даже если несколько позиций совпадают по
названию.
"""

import uuid
from datetime import datetime

from openpyxl import Workbook, load_workbook

from modules.paths import WEIGHT_LOG_FILE

COLUMNS = (
    "ID",
    "Дата",
    "Наименование",
    "Кол-во ящиков",
    "Средний вес ящика, кг",
    "Грязный вес, кг",
    "Чистый вес, кг",
    "Заказ",
    "Маршрут",
    "Этап",
    "Магазин",
)

STAGE_LOADING = "Загрузка"
STAGE_UNLOADING = "Выгрузка"
STAGE_STORE_SHIPMENT = "Отгрузка с магазинов"
STAGES = (STAGE_LOADING, STAGE_UNLOADING, STAGE_STORE_SHIPMENT)


def _open_workbook():
    if WEIGHT_LOG_FILE.exists():
        return load_workbook(WEIGHT_LOG_FILE)

    workbook = Workbook()
    workbook.active.title = "Вес"
    workbook.active.append(COLUMNS)
    return workbook


def _save_workbook(workbook) -> None:
    WEIGHT_LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(WEIGHT_LOG_FILE)


def load_weight_rows() -> list[dict]:
    """Записи из data/weight_log.xlsx в порядке добавления."""

    if not WEIGHT_LOG_FILE.exists():
        return []

    sheet = load_workbook(WEIGHT_LOG_FILE).active
    rows = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_id, date, name, box_count, avg_weight, exact_weight, total = row[:7]
        order_file, route, stage, store = (row[7:11] + (None, None, None, None))[:4]

        if row_id is None:
            continue

        rows.append(
            {
                "id": row_id,
                "date": date or "",
                "name": name or "",
                "box_count": box_count or 0,
                "avg_weight": avg_weight or 0,
                "exact_weight": exact_weight,
                "total": total or 0,
                "order_file": order_file or "",
                "route": route or "",
                "stage": stage or "",
                "store": store or "",
            }
        )

    return rows


def add_weight_row(
    name: str,
    box_count: float,
    avg_weight: float,
    exact_weight: float | None,
    order_file: str = "",
    route: str = "",
    stage: str = "",
    store: str = "",
) -> dict:
    """Добавляет строку и возвращает её.

    Итог (чистый вес) — если позицию взвесили вместе с ящиками (exact_weight
    — грязный вес), из него вычитается вес самих ящиков (кол-во ящиков ×
    средний вес ящика); иначе итог — это оценка, кол-во ящиков, умноженное
    на средний вес ящика. order_file и route — необязательная привязка к
    обработанному заказу и маршруту, для группировки записей. stage — один
    из STAGES (загрузка/выгрузка/отгрузка с магазинов), store — магазин,
    актуален только для этапа отгрузки с магазинов.
    """

    tare = box_count * avg_weight
    entry = {
        "id": uuid.uuid4().hex,
        "date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "name": name,
        "box_count": box_count,
        "avg_weight": avg_weight,
        "exact_weight": exact_weight,
        "total": (exact_weight - tare) if exact_weight is not None else tare,
        "order_file": order_file,
        "route": route,
        "stage": stage,
        "store": store,
    }

    workbook = _open_workbook()
    workbook.active.append(
        [
            entry["id"],
            entry["date"],
            entry["name"],
            entry["box_count"],
            entry["avg_weight"],
            entry["exact_weight"],
            entry["total"],
            entry["order_file"],
            entry["route"],
            entry["stage"],
            entry["store"],
        ]
    )
    _save_workbook(workbook)

    return entry


def delete_weight_row(row_id: str) -> None:
    """Удаляет строку по id, если она есть в книге."""

    if not WEIGHT_LOG_FILE.exists():
        return

    workbook = _open_workbook()
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2):
        if row[0].value == row_id:
            sheet.delete_rows(row[0].row)
            _save_workbook(workbook)
            return


def update_weight_row(
    row_id: str,
    name: str,
    box_count: float,
    avg_weight: float,
    exact_weight: float | None,
    order_file: str = "",
    route: str = "",
    stage: str = "",
    store: str = "",
) -> dict | None:
    """Обновляет существующую строку по id и возвращает её новую версию.

    Дата и id исходной записи сохраняются. Возвращает None, если строка не
    найдена. Пишет через sheet.cell(...) вместо ячеек из iter_rows — так
    правка не падает на книге, сохранённой до появления колонок Заказ,
    Маршрут, Этап и Магазин (в ней короче строк, чем ожидает текущая схема).
    """

    if not WEIGHT_LOG_FILE.exists():
        return None

    workbook = _open_workbook()
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2):
        if row[0].value != row_id:
            continue

        row_number = row[0].row
        tare = box_count * avg_weight
        total = (exact_weight - tare) if exact_weight is not None else tare
        date = sheet.cell(row=row_number, column=2).value

        sheet.cell(row=row_number, column=3, value=name)
        sheet.cell(row=row_number, column=4, value=box_count)
        sheet.cell(row=row_number, column=5, value=avg_weight)
        sheet.cell(row=row_number, column=6, value=exact_weight)
        sheet.cell(row=row_number, column=7, value=total)
        sheet.cell(row=row_number, column=8, value=order_file)
        sheet.cell(row=row_number, column=9, value=route)
        sheet.cell(row=row_number, column=10, value=stage)
        sheet.cell(row=row_number, column=11, value=store)
        _save_workbook(workbook)

        return {
            "id": row_id,
            "date": date or "",
            "name": name,
            "box_count": box_count,
            "avg_weight": avg_weight,
            "exact_weight": exact_weight,
            "total": total,
            "order_file": order_file,
            "route": route,
            "stage": stage,
            "store": store,
        }

    return None


def known_names_for_order(order_file: str) -> list[str]:
    """Наименования, которые уже вводили для этого заказа (любой этап).

    Подсказка для массового ввода: у заказа обычно повторяется один и тот же
    набор позиций от раза к разу. От самых свежих к самым старым, без
    повторов.
    """

    order_file = order_file.strip()
    if not order_file:
        return []

    seen: set[str] = set()
    names: list[str] = []
    for row in reversed(load_weight_rows()):
        if row["order_file"] != order_file:
            continue
        name = row["name"].strip()
        key = name.lower()
        if not name or key in seen:
            continue
        seen.add(key)
        names.append(name)

    return names


def last_avg_weight_for(name: str) -> float | None:
    """Средний вес ящика из самой последней записи с таким наименованием.

    Подсказка для формы: одно и то же наименование часто взвешивают
    регулярно, и вспоминать вес ящика каждый раз заново неудобно.
    """

    normalized = name.strip().lower()
    if not normalized:
        return None

    match = None
    for row in load_weight_rows():
        if row["name"].strip().lower() == normalized:
            match = row

    return match["avg_weight"] if match else None
