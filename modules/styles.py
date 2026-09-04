from openpyxl.styles import PatternFill


# №1 Зеленый
green_fill = PatternFill(
    fill_type="solid",
    start_color="00FF00",
    end_color="00FF00"
)


# №2 Желтый
yellow_fill = PatternFill(
    fill_type="solid",
    start_color="FFFF00",
    end_color="FFFF00"
)


# №3 Голубой
blue_fill = PatternFill(
    fill_type="solid",
    start_color="00B0F0",
    end_color="00B0F0"
)


# №4 Фиолетовый
purple_fill = PatternFill(
    fill_type="solid",
    start_color="800080",
    end_color="800080"
)


# Конфликт
conflict_fill = PatternFill(
    fill_type="solid",
    start_color="FF0000",
    end_color="FF0000"
)


# Заливки для build_groups по числу маршрутов (до config.MAX_ROUTES).
# Первые 4 совпадают с green_fill/yellow_fill/blue_fill/purple_fill выше —
# существующие маршруты 1-4 не перекрашиваются при добавлении новых.
ROUTE_FILL_COLORS = (
    "00FF00",  # №1 зелёный
    "FFFF00",  # №2 жёлтый
    "00B0F0",  # №3 голубой
    "800080",  # №4 фиолетовый
    "FFA500",  # №5 оранжевый
    "FF69B4",  # №6 розовый
    "20B2AA",  # №7 бирюзовый
    "8B4513",  # №8 коричневый
)


def fills_for(count: int) -> list[PatternFill]:
    """count заливок для build_groups, в порядке route_1..route_N."""

    return [
        PatternFill(fill_type="solid", start_color=color, end_color=color)
        for color in ROUTE_FILL_COLORS[:count]
    ]

