import re
from copy import copy

from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Font

SUM_HEADER = "Сумма"
DEFAULT_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def create_route_sheets(wb, ws, groups):
    """
    Для каждой группы (маршрута) создаёт отдельный лист-копию исходного
    листа, оставляя только столбцы, относящиеся к этому маршруту,
    и добавляет колонку "Сумма" с формулой SUM по строке.
    """

    for group in groups:

        route_name = group["name"]
        stores = group["names"]

        # копируем весь лист
        new_ws = wb.copy_worksheet(ws)
        new_ws.title = route_name

        delete_columns = _find_columns_to_delete(new_ws, stores)

        # удаляем справа налево, чтобы номера столбцов не сбились
        for col in sorted(delete_columns, reverse=True):
            new_ws.delete_cols(col)


def _find_columns_to_delete(ws, stores):
    """Определяет столбцы, не относящиеся к данному маршруту (кроме 1-го)."""

    delete_columns = []

    # начинаем со 2-го столбца - первый оставляем всегда
    for col in range(2, ws.max_column + 1):

        header = str(ws.cell(1, col).value or "").lower()
        is_route_column = False

        for store in stores:

            pattern = r"\b" + re.escape(store.lower()) + r"\b"

            if re.search(pattern, header):
                is_route_column = True
                break

        if not is_route_column:
            delete_columns.append(col)

    return delete_columns


def add_sum_column_to_all_sheets(wb):

    route_totals = {}

    for ws in wb.worksheets:

        total = _add_sum_column(ws)

        if ws.title.startswith("Маршрут №"):
            route_totals[ws.title] = total

    return route_totals

def _add_sum_column(ws):
    """Adds or updates the row sum column as the second worksheet column."""

    if not _has_sum_column(ws):
        ws.insert_cols(2)

    header_cell = ws.cell(row=1, column=2, value=SUM_HEADER)

    last_col = _last_used_column(ws)
    _delete_rows_without_orders(ws, last_col)
    last_col = _last_used_column(ws)

    header_cell.border = _border_for_row(ws, 1, last_col)

    route_total = 0

    for row in range(2, ws.max_row + 1):

        row_total = 0

        for col in range(3, last_col + 1):

            value = ws.cell(row=row, column=col).value

            if isinstance(value, (int, float)):
                row_total += value

        route_total += row_total

        sum_cell = ws.cell(
            row=row,
            column=2,
            value=f"=SUM(C{row}:{get_column_letter(last_col)}{row})",
        )
        sum_cell.border = _border_for_row(ws, row, last_col)
    # Запоминаем последнюю строку с товаром
    data_last_row = ws.max_row

    # Итог по листу через 3 пустые строки
    total_row = data_last_row + 4

    ws.cell(total_row, 1, "ИТОГО")

    ws.cell(
        total_row,
        2,
        f"=SUM(B2:B{data_last_row})"
    ) 
    ws.cell(total_row, 1).font = Font(bold=True)
    ws.cell(total_row, 2).font = Font(bold=True) 
    ws.column_dimensions["B"].width = 10

    return route_total


def _delete_rows_without_orders(ws, last_col):
    """Deletes rows that have no order values in store columns."""

    rows_to_delete = []

    for row in range(2, ws.max_row + 1):
        if not _row_has_data(ws, row, 3, last_col):
            rows_to_delete.append(row)

    for row in reversed(rows_to_delete):
        ws.delete_rows(row)


def _row_has_data(ws, row, start_col, end_col):
    """Checks whether a row has at least one non-empty cell in a column range."""

    for col in range(start_col, end_col + 1):
        if ws.cell(row=row, column=col).value not in (None, ""):
            return True

    return False


def _border_for_row(ws, row, last_col):
    """Returns an existing row border, or a plain fallback border."""

    for col in [1, *range(3, last_col + 1)]:
        border = ws.cell(row=row, column=col).border

        if _has_visible_border(border):
            return copy(border)

    return copy(DEFAULT_BORDER)


def _has_visible_border(border):
    """Checks whether a border has at least one visible side."""

    return any(
        side is not None and side.style is not None
        for side in [
            border.left,
            border.right,
            border.top,
            border.bottom,
            border.diagonal,
            border.vertical,
            border.horizontal,
        ]
    )


def _has_sum_column(ws):
    """Checks whether column B already contains the sum column."""

    header = str(ws.cell(row=1, column=2).value or "").strip().lower()

    if header == SUM_HEADER.lower():
        return True

    for row in range(2, ws.max_row + 1):
        value = ws.cell(row=row, column=2).value

        if isinstance(value, str) and value.strip().upper().startswith("=SUM("):
            return True

    return False


def _last_used_column(ws):
    """Returns the rightmost column that has a non-empty cell."""

    for col in range(ws.max_column, 2, -1):
        for row in range(1, ws.max_row + 1):
            if ws.cell(row=row, column=col).value not in (None, ""):
                return col

    return 2
